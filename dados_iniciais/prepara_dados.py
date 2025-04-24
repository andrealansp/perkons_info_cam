import re
from pathlib import Path

import pandas as pd
from openpyxl import load_workbook


def formata_ip(ip: str) -> str:
    if ip.find(".", 0, len(ip)) == -1:
        ip = re.sub(r'(\d{2,3})(\d{3})(\d{3})(\d{3})', r"\1.\2.\3.\4", ip)
        print(ip)
    return ip

p = Path("cameras.xlsx")

wb = load_workbook(p.resolve())
ws = wb["status"]

for row in ws.iter_rows(min_row=2):
    try:
        ip = formata_ip(str(row[3].value))
        row[3].value = ip
    except IndexError:
        continue

wb.save("cameras2.xlsx")
